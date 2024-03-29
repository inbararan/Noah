import os
import openpyxl
from reader import Candidate, ExerciseType
from typing import List, Dict
import identify
import paths


REQUIRED_ATTRIBUTES = ["learning_ability", "personal", "interpersonal", "leader", "summary"]


def _extract_evaluators(candidates: List[Candidate]) -> List[str]:
    """
    Accepts a list of candidates, and returns a sorted list of
    evaluators which have evaluated at least one of the candidates
    """
    evaluators = set()
    for candidate in candidates:
        for evaluation in candidate.evaluations:
            evaluators.add(evaluation.evaluator_name)
    if len(evaluators) > 4:
        print(f"WARNING: Too much evaluators for single team: {evaluators}")
    return sorted(evaluators)


def _write_file_for_team(team_name: str, candidates: List[Candidate], exercise_name):
    # Plan is: load template, modify it, and save to a new file
    wb = openpyxl.load_workbook(paths.HAZANOTOMAT_TEMPLATE)

    for sheet in wb.worksheets:
        sheet.cell(1, 1, exercise_name)

    evaluators = _extract_evaluators(candidates)

    # Fill in evaluator names
    for evaluator_index, evaluator in enumerate(evaluators):
        col_index = 2 + evaluator_index * 2
        for sheet in wb.worksheets:
            sheet.cell(1, col_index, evaluator)

    # Fill in name and evaluation for each candidate
    for candidate_index, candidate in enumerate(candidates):
        row_index = candidate_index + 2

        candidate_name = identify.get_private_name(candidate.uuid)
        for sheet in wb.worksheets:
            sheet.cell(row_index, 1, candidate_name)
        for evaluation in candidate.evaluations:
            evaluator_index = evaluators.index(evaluation.evaluator_name)
            col_index_num = 2 + evaluator_index * 2
            col_index_text = col_index_num + 1

            for i, attr_name in enumerate(REQUIRED_ATTRIBUTES):
                attribute = getattr(evaluation, attr_name)
                wb.worksheets[i].cell(row_index, col_index_num, attribute.num)
                wb.worksheets[i].cell(row_index, col_index_text, attribute.text)

    os.makedirs(os.path.join(paths.HAZANOTOMAT_OUTPUT, exercise_name), exist_ok=True)  # Make sure directory exists
    wb.save(os.path.join(paths.HAZANOTOMAT_OUTPUT, exercise_name, f"צוות {team_name}.xlsx"))


def _filter_evaluations(candidate: Candidate, exercise_name: str) -> Candidate:
    """
    Accepts a candidate and an exercise name.
    Returns a similar candidate with evaluations which are only of the specified exercise.
    """
    return Candidate(
        candidate.uuid,
        [evaluation for evaluation in candidate.evaluations if evaluation.exercise_name == exercise_name]
    )


def write(teams: Dict[str, List[Candidate]]):
    print("Writing...")
    for team_name, candidates_in_team in teams.items():
        for exercise in ExerciseType.all():
            candidates = [_filter_evaluations(candidate, exercise.name) for candidate in candidates_in_team]
            _write_file_for_team(team_name, candidates, exercise.name)
